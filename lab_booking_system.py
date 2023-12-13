import tkinter as tk
from tkinter import ttk
from tkinter import simpledialog, messagebox
from tkinter import simpledialog, Label, Entry, Toplevel, Button
from tkcalendar import Calendar
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

class LabBookingSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Lab Booking System")

        self.user_info_frame()
        self.lab_selection_frame()
        self.date_time_frame()
        self.terms_conditions_frame()
        self.action_buttons_frame()

        self.tree = ttk.Treeview(self.root, columns=("ID", "First Name", "Last Name", "Matric No", "Faculty", "Programme", "Semester", "Lab", "Date", "Time"))
        self.tree.heading("#0", text="")
        self.tree.heading("ID", text="ID")
        self.tree.heading("First Name", text="First Name")
        self.tree.heading("Last Name", text="Last Name")
        self.tree.heading("Matric No", text="Matric No")
        self.tree.heading("Faculty", text="Faculty")
        self.tree.heading("Programme", text="Programme")
        self.tree.heading("Semester", text="Semester")
        self.tree.heading("Lab", text="Lab")
        self.tree.heading("Date", text="Date")
        self.tree.heading("Time", text="Time")

        self.tree.grid(row=6, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

        # Configure the grid row and column weights for resizing
        self.root.grid_rowconfigure(6, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_columnconfigure(2, weight=1)

    def user_info_frame(self):
        frame = ttk.LabelFrame(self.root, text="User Information")
        frame.grid(row=0, column=0, padx=10, pady=5, sticky="nsew", columnspan=3)

        ttk.Label(frame, text="First Name:").grid(row=0, column=0, padx=5, pady=5)
        self.first_name_entry = ttk.Entry(frame)
        self.first_name_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(frame, text="Last Name:").grid(row=0, column=2, padx=5, pady=5)
        self.last_name_entry = ttk.Entry(frame)
        self.last_name_entry.grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(frame, text="Matric No:").grid(row=0, column=4, padx=5, pady=5)
        self.matric_no_entry = ttk.Entry(frame)
        self.matric_no_entry.grid(row=0, column=5, padx=5, pady=5)

        ttk.Label(frame, text="Faculty:").grid(row=1, column=0, padx=5, pady=5)
        self.faculty_combobox = ttk.Combobox(frame, values=["CDIM", "IM", "AM"])
        self.faculty_combobox.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(frame, text="Programme:").grid(row=1, column=2, padx=5, pady=5)
        self.programme_combobox = ttk.Combobox(frame, values=["Diploma", "Degree"])
        self.programme_combobox.grid(row=1, column=3, padx=5, pady=5)

        ttk.Label(frame, text="Semester:").grid(row=1, column=4, padx=5, pady=5)
        self.semester_spinbox = ttk.Spinbox(frame, from_=1, to=6)
        self.semester_spinbox.grid(row=1, column=5, padx=5, pady=5)

    def lab_selection_frame(self):
        frame = ttk.LabelFrame(self.root, text="Lab Selection")
        frame.grid(row=1, column=0, padx=10, pady=5, sticky="nsew", columnspan=3)

        ttk.Label(frame, text="Lab:").grid(row=0, column=0, padx=5, pady=5)
        self.lab_combobox = ttk.Combobox(frame, values=["Lab A", "Lab B", "Lab C"])
        self.lab_combobox.grid(row=0, column=1, padx=5, pady=5)

    def date_time_frame(self):
        frame = ttk.LabelFrame(self.root, text="Date and Time Selection")
        frame.grid(row=2, column=0, columnspan=3, padx=10, pady=5, sticky="nsew")

        ttk.Label(frame, text="Date:").grid(row=0, column=0, padx=5, pady=5)
        self.date_entry = ttk.Entry(frame)
        self.date_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Button(frame, text="Pick Date", command=self.pick_date).grid(row=0, column=2, padx=5, pady=5)

        ttk.Label(frame, text="Time:").grid(row=1, column=0, padx=5, pady=5)
        self.time_combobox = ttk.Combobox(frame, values=["8AM - 10AM", "10AM - 12PM", "12PM - 2PM", "2PM - 4PM", "4PM - 6PM", "6PM - 8PM", "8PM - 10PM"])
        self.time_combobox.grid(row=1, column=1, padx=5, pady=5)

    def terms_conditions_frame(self):
        frame = ttk.LabelFrame(self.root, text="Terms and Conditions")
        frame.grid(row=3, column=0, columnspan=3, padx=10, pady=5, sticky="nsew")

        self.agree_var = tk.IntVar()
        self.agree_checkbox = ttk.Checkbutton(frame, text="I agree to the terms and conditions", variable=self.agree_var)
        self.agree_checkbox.grid(row=0, column=0, padx=5, pady=5)

    def action_buttons_frame(self):
        frame = ttk.LabelFrame(self.root, text="Actions")
        frame.grid(row=4, column=0, columnspan=3, padx=10, pady=5, sticky="nsew")

        ttk.Button(frame, text="Book Lab", command=self.book_lab).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(frame, text="Edit Booking", command=self.edit_booking).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(frame, text="Delete Booking", command=self.delete_booking).grid(row=0, column=2, padx=5, pady=5)

    def pick_date(self):
        date_str = Calendar(self.root).get_date()
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, date_str)

    def validate_input(self):
        if not all([self.first_name_entry.get(), self.last_name_entry.get(), self.matric_no_entry.get(),
                    self.faculty_combobox.get(), self.programme_combobox.get(), self.semester_spinbox.get(),
                    self.lab_combobox.get(), self.date_entry.get(), self.time_combobox.get()]):
            messagebox.showerror("Error", "Please fill in all details.")
            return False
        elif not self.agree_checkbox.instate(['selected']):
            messagebox.showerror("Error", "Please agree to the terms and conditions.")
            return False
        return True

    def book_lab(self):
        if not self.validate_input():
            return

        # Get data from the user input
        data = [
            len(self.tree.get_children()) + 1,
            self.first_name_entry.get(),
            self.last_name_entry.get(),
            self.matric_no_entry.get(),
            self.faculty_combobox.get(),
            self.programme_combobox.get(),
            self.semester_spinbox.get(),
            self.lab_combobox.get(),
            self.date_entry.get(),
            self.time_combobox.get(),
            "Yes" if self.agree_checkbox.instate(['selected']) else "No"
        ]

        # Add data to the treeview
        self.tree.insert("", "end", values=data)

        # Save data to Excel sheet
        self.save_to_excel(data)

        # Clear input fields
        self.clear_inputs()
        self.agree_var.set(0)

        # Display a success message
        messagebox.showinfo("Booking Successful", "Your lab booking is successful!")

    def edit_booking(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a booking to edit.")
            return

        # Retrieve the selected item's data
        selected_data = self.tree.item(selected_item, "values")

        # Open a dialog for editing
        edited_data = self.edit_dialog(selected_data)
        if edited_data:
            # Update the treeview with the edited data
            self.tree.item(selected_item, values=edited_data)

    def edit_dialog(self, data):
        # Exclude the first element (ID) when displaying the data
        data_without_id = data[1:]

        # Create a Toplevel window for editing
        edit_window = Toplevel(self.root)
        edit_window.title("Edit Booking")

        # Create entry widgets for each field
        labels = ["First Name", "Last Name", "Matric No", "Faculty", "Programme", "Semester", "Lab", "Date", "Time", "Agreed"]
        entry_widgets = []

        for i, label in enumerate(labels):
            Label(edit_window, text=label).grid(row=i, column=0, padx=5, pady=5)
            entry_var = tk.StringVar()
            entry_var.set(data_without_id[i])  # Set the initial value
            Entry(edit_window, textvariable=entry_var).grid(row=i, column=1, padx=5, pady=5)
            entry_widgets.append(entry_var)

        # Initialize edited_data before the ok_button_clicked function
        edited_data = [data[0]]

        # Function to handle the OK button click
        def ok_button_clicked():
            nonlocal edited_data, entry_widgets
            edited_data += [entry_var.get() for entry_var in entry_widgets]
            edit_window.destroy()  # Close the editing window
            # You can perform additional validation or processing here before returning the edited data

        # Create an OK button
        ok_button = Button(edit_window, text="OK", command=ok_button_clicked)
        ok_button.grid(row=len(labels), column=0, columnspan=2, pady=10)

        # Run the edit window in a modal state
        edit_window.transient(self.root)
        edit_window.wait_window()

        # The function will return the edited data
        return edited_data

    
    def delete_booking(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a booking to delete.")
            return

        # Delete selected item from the treeview
        self.tree.delete(selected_item)

        # Update the Excel sheet after deletion
        self.update_excel_after_deletion(selected_item)

    def save_to_excel(self, data):
        workbook_path = "lab_bookings.xlsx"
        try:
            workbook = load_workbook(workbook_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "First Name", "Last Name", "Matric No", "Faculty", "Programme", "Semester", "Lab", "Date", "Time", "Agreed"])

        sheet.append(data)
        workbook.save(workbook_path)

    def update_excel_after_deletion(self, selected_item):
        workbook_path = "lab_bookings.xlsx"
        workbook = load_workbook(workbook_path)
        sheet = workbook.active

        # Assume ID is stored in the first column
        selected_id = int(self.tree.item(selected_item, "values")[0])

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == selected_id:
                sheet.delete_rows(row[0].row)
                break

        workbook.save(workbook_path)

    def clear_inputs(self):
        self.first_name_entry.delete(0, tk.END)
        self.last_name_entry.delete(0, tk.END)
        self.matric_no_entry.delete(0, tk.END)
        self.faculty_combobox.set("")
        self.programme_combobox.set("")
        self.semester_spinbox.set("")
        self.lab_combobox.set("")
        self.date_entry.delete(0, tk.END)
        self.time_combobox.set("")
        self.agree_checkbox.deselect()

if __name__ == "__main__":
    root = tk.Tk()
    app = LabBookingSystem(root)
    root.mainloop()
